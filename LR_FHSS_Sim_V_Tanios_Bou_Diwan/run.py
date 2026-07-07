# -*- coding: utf-8 -*-
from lrfhss.lrfhss_core import *
from lrfhss.acrda import BaseACRDA
from lrfhss.settings import Settings

import simpy
import numpy as np
from time import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# Paramètres
# ============================================================
NODE_COUNTS     = [50000]
SIC_LIMITS      = [None]
GAMMAS          = [1]
SEEDS           = list(range(1))
SIMULATION_TIME = 3600
PAYLOAD_SIZE    = 10
CODE            = '1/3'
OBW             = 35
BASE            = 'acrda'   # <-- change ici en 'core' si tu veux tester sans ACRDA

DISTRIBUTIONS = [
    #('h2', [(2, 1)]),
    ('h3', [(3, 1)]),
    #('h2=h3=0.5',       [(2, 0.5), (3, 0.5)]),  # 50% h=2, 50% h=3
]

# ============================================================
# Simulation générique multi-groupes
# ============================================================
def run_sim_groups(groups_spec, total_nodes, seed=0, sic_limit=None, gamma=1.0):
    random.seed(seed)
    np.random.seed(seed)
    env = simpy.Environment()

    settings_list = []
    for h, prop in groups_spec:
        s = Settings(number_nodes=total_nodes, simulation_time=SIMULATION_TIME,
                     payload_size=PAYLOAD_SIZE, headers=h, code=CODE,
                     obw=OBW, base=BASE, sic_limit=sic_limit, gamma=gamma)
        settings_list.append((s, prop))

    s0 = settings_list[0][0]

    # --- Branchement selon le type de base (correction du bug) ---
    if s0.base == 'acrda':
        avg_toa = np.mean([s.time_on_air for s, _ in settings_list])
        bs = BaseACRDA(OBW, s0.window_size, s0.window_step, avg_toa, s0.threshold, sic_limit, gamma, seed)
        env.process(bs.sic_window(env))
    else:
        bs = Base(OBW, s0.threshold)
    # ---------------------------------------------------------------

    nodes = []
    remaining = total_nodes
    for i, (s, prop) in enumerate(settings_list):
        n = int(prop * total_nodes) if i < len(settings_list) - 1 else remaining
        remaining -= n
        for _ in range(n):
            node = Node(s.obw, s.headers, s.payloads,
                        s.header_duration, s.payload_duration,
                        s.transceiver_wait, s.traffic_generator)
            bs.add_node(node.id)
            nodes.append(node)
            env.process(node.transmit(env, bs))

    env.run(until=SIMULATION_TIME)

    success     = sum(bs.packets_received.values())
    transmitted = sum(n.transmitted for n in nodes)

    return success / transmitted if transmitted > 0 else 1.0

# ============================================================
# Helpers Excel
# ============================================================
THIN = Side(style='thin', color='000000')

def _border(*sides):
    return Border(**{s: THIN for s in sides})

def build_table(ws, label, results, node_counts, gammas, start_row):
    ws.cell(start_row, 1, label).font = Font(bold=True, size=11)
    start_row += 1

    gamma_start_col = 2
    cell = ws.cell(start_row, gamma_start_col, "Gamma")
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    ws.merge_cells(start_row=start_row, start_column=gamma_start_col,
                   end_row=start_row, end_column=gamma_start_col + len(gammas) - 1)
    start_row += 1

    c = ws.cell(start_row, 1, "Nodes")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    c.fill = PatternFill('solid', fgColor='FFFF00')
    c.border = _border('left', 'right', 'top', 'bottom')

    for j, g in enumerate(gammas):
        c = ws.cell(start_row, gamma_start_col + j, g)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = _border('left', 'right', 'top', 'bottom')
    start_row += 1

    for n in node_counts:
        c = ws.cell(start_row, 1, f"{n // 1000}k")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill('solid', fgColor='FFFF00')
        c.border = _border('left', 'right', 'top', 'bottom')
        for j, g in enumerate(gammas):
            mean, _ = results.get((n, g), (None, None))
            c = ws.cell(start_row, gamma_start_col + j)
            if mean is not None:
                c.value = round(mean, 2)
            c.alignment = Alignment(horizontal='center')
            c.number_format = '0.00'
            c.border = _border('left', 'right', 'top', 'bottom')
        start_row += 1

    return start_row + 1

def save_to_excel(all_results, node_counts, gammas, distributions, sic_limits, filename="lrfhss_results_pas_mod_halifax.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    for label, _ in distributions:
        for sic in sic_limits:
            sic_label = str(sic) if sic is not None else 'None'
            ws = wb.create_sheet(title=f"{label}_SIC={sic_label}")
            ws.column_dimensions['A'].width = 9
            for j in range(len(gammas)):
                ws.column_dimensions[get_column_letter(j + 2)].width = 8

            results = {
                (n, g): all_results[(label, sic, n, g)]
                for n in node_counts for g in gammas
                if (label, sic, n, g) in all_results
            }
            title = f"{label.upper()}, Gamma, SIC={sic_label}, moyenne sur {len(SEEDS)} seeds"
            build_table(ws, title, results, node_counts, gammas, start_row=1)

    wb.save(filename)
    print(f"\n✅ Résultats sauvegardés dans : {filename}")

# ============================================================
# Main
# ============================================================
if __name__ == "__main__":

    start_time = time()
    all_results = {}  # {(label, sic, n, g): (mean, std)}

    total_jobs = len(DISTRIBUTIONS) * len(NODE_COUNTS) * len(SIC_LIMITS) * len(GAMMAS)
    print(f"Total jobs : {total_jobs} (moyenne sur {len(SEEDS)} seeds chacun)")
    print(f"{'Dist':<20} {'Nodes':>8} {'SIC':>6} {'Gamma':>6} {'Succes moy (%)':>15} {'Std (%)':>8}")
    print("-" * 70)

    for label, groups_spec in DISTRIBUTIONS:
        for sic in SIC_LIMITS:
            for g in GAMMAS:
                for n in NODE_COUNTS:
                    rates = [
                        run_sim_groups(groups_spec, total_nodes=n // 8,
                                       seed=s, sic_limit=sic, gamma=g)
                        for s in SEEDS
                    ]
                    mean = np.mean(rates) * 100
                    std  = np.std(rates)  * 100

                    all_results[(label, sic, n, g)] = (mean, std)

                    sic_str = str(sic) if sic is not None else 'None'
                    print(f"{label:<20} {n:>8} {sic_str:>6} {g:>6.2f} {mean:>15.2f} {std:>8.2f}",
                          flush=True)
'''
    save_to_excel(
        all_results=all_results,
        node_counts=NODE_COUNTS,
        gammas=GAMMAS,
        distributions=DISTRIBUTIONS,
        sic_limits=SIC_LIMITS,
        filename="lrfhss_results_bis.xlsx"
    )

    elapsed = time() - start_time
    print(f"\nElapsed time: {elapsed:.2f} seconds")
'''
